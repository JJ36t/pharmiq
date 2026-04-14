"""PharmIQ — Excel Reports"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from database.crud import get_all_drugs, get_all_sales

def generate_drugs_excel(session,filename="drugs_report.xlsx"):
    drugs=get_all_drugs(session); wb=openpyxl.Workbook(); ws=wb.active; ws.title="Drugs"
    headers=["ID","Trade Name","Scientific Name","Category","Qty","Min Qty","Price (IQD)","Expiry","Status"]
    ws.append(headers)
    for col in range(1,len(headers)+1):
        cell=ws.cell(row=1,column=col); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill(start_color="2E6DA4",end_color="2E6DA4",fill_type="solid")
        cell.alignment=Alignment(horizontal="center")
    for d in drugs:
        status=""
        if d.is_expired: status="❌ منتهي الصلاحية"
        elif d.is_low_stock: status="⚠️ مخزون قليل"
        elif d.is_expiring_soon: status="🔴 ينتهي قريباً"
        ws.append([d.id,d.trade_name,d.scientific_name or "",d.category or "",d.quantity,d.min_quantity,d.price,str(d.expiry_date) if d.expiry_date else "",status])
        if d.is_expired or d.is_low_stock or d.is_expiring_soon:
            for col in range(1,len(headers)+1):
                ws.cell(row=ws.max_row,column=col).fill=PatternFill(start_color="FFE0E0",end_color="FFE0E0",fill_type="solid")
    for col in ws.columns:
        ml=max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width=ml+4
    wb.save(filename); return filename

def generate_sales_excel(session,filename="sales_report.xlsx"):
    sales=get_all_sales(session); wb=openpyxl.Workbook(); ws=wb.active; ws.title="Sales"
    headers=["ID","Drug Name","Qty Sold","Total (IQD)","Sale Date","Pharmacist"]
    ws.append(headers)
    for col in range(1,len(headers)+1):
        cell=ws.cell(row=1,column=col); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill(start_color="27AE60",end_color="27AE60",fill_type="solid")
        cell.alignment=Alignment(horizontal="center")
    total=0
    for s in sales:
        dn=s.drug.trade_name if s.drug else "N/A"
        sd=s.sale_date.strftime("%Y-%m-%d %H:%M") if s.sale_date else ""
        ph=s.pharmacist.username if s.pharmacist else "—"
        ws.append([s.id,dn,s.quantity_sold,s.total_price,sd,ph]); total+=s.total_price
    ws.append([]); ws.append(["","","Total Revenue:",total,"",""])
    lr=ws.max_row; ws.cell(row=lr,column=3).font=Font(bold=True); ws.cell(row=lr,column=4).font=Font(bold=True)
    for col in ws.columns:
        ml=max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width=ml+4
    wb.save(filename); return filename
